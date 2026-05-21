from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from .models import (Requester, Status, Platform, Request,
                     Approval, Attachment, Publication, ProcessingLog)
from .serializers import (RequesterSerializer, StatusSerializer,
                          PlatformSerializer, RequestSerializer,
                          ApprovalSerializer, AttachmentSerializer,
                          PublicationSerializer, ProcessingLogSerializer)
from logs.utils import log_action
from notifications.utils import send_notification
from django.contrib.auth import get_user_model
from users.permissions import IsStaffOrAdmin, IsOwnerOrStaff
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from io import BytesIO

User = get_user_model()


class RequesterViewSet(viewsets.ModelViewSet):
    queryset = Requester.objects.all()
    serializer_class = RequesterSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        queryset = Requester.objects.all()
        if not user.is_authenticated:
            return queryset.none()
        # Citizens only see their own information
        if hasattr(user, 'role') and user.role and user.role.role_name == 'citizen':
            return queryset.filter(email=user.email)
        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'CREATE REQUESTER',
                       f'Added requester: {instance.agency_name}')

    def perform_update(self, serializer):
        instance = serializer.save()
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'UPDATE REQUESTER',
                       f'Updated requester: {instance.agency_name}')

    def perform_destroy(self, instance):
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'DELETE REQUESTER',
                       f'Deleted requester: {instance.agency_name}')
        instance.delete()


class StatusViewSet(viewsets.ModelViewSet):
    queryset = Status.objects.all()
    serializer_class = StatusSerializer
    permission_classes = [permissions.IsAuthenticated]


class PlatformViewSet(viewsets.ModelViewSet):
    queryset = Platform.objects.all()
    serializer_class = PlatformSerializer
    permission_classes = [permissions.IsAuthenticated]


class RequestViewSet(viewsets.ModelViewSet):
    queryset = Request.objects.all()
    serializer_class = RequestSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [IsOwnerOrStaff()]

    def get_queryset(self):
        queryset = Request.objects.all().order_by('-submitted_at')
        user = self.request.user
        if not user.is_authenticated:
            return queryset.none()
        request_type = self.request.query_params.get('request_type')
        status = self.request.query_params.get('status')
        search = self.request.query_params.get('search')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        # Citizens only see their own requests
        if hasattr(user, 'role') and user.role and user.role.role_name == 'citizen':
            queryset = queryset.filter(requester__email=user.email)

        if request_type:
            queryset = queryset.filter(request_type=request_type)
        if status:
            queryset = queryset.filter(status__status_name=status)
        if search:
            queryset = queryset.filter(details__icontains=search)
        if date_from:
            queryset = queryset.filter(submitted_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(submitted_at__date__lte=date_to)

        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'CREATE REQUEST',
                       f'Submitted request #{instance.request_id}: {instance.get_request_type_display()}')
            send_notification(
                user=self.request.user,
                message=f'Your request #{instance.request_id} ({instance.get_request_type_display()}) has been successfully submitted and is now pending review.',
                notification_type='request_submitted'
            )

    def perform_update(self, serializer):
        instance = serializer.save()
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'UPDATE REQUEST',
                       f'Updated request #{instance.request_id}')
            send_notification(
                user=self.request.user,
                message=f'Your request #{instance.request_id} ({instance.get_request_type_display()}) has been updated.',
                notification_type='request_updated'
            )

    def perform_destroy(self, instance):
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'DELETE REQUEST',
                       f'Deleted request #{instance.request_id}')
        instance.delete()

    @action(detail=False, methods=['get'], permission_classes=[IsStaffOrAdmin])
    def analytics(self, request):
        """
        Returns a breakdown of requests by Barangay.
        """
        stats = Request.objects.values('requester__barangay').annotate(
            total=Count('request_id')
        ).order_by('-total')

        # Format the data for the frontend
        formatted_stats = []
        # Create a dictionary of current choices for lookup
        choices_dict = dict(Requester._meta.get_field('barangay').choices)

        for entry in stats:
            b_code = entry['requester__barangay']
            formatted_stats.append({
                'barangay': choices_dict.get(b_code, b_code or 'Unknown'),
                'count': entry['total']
            })

        return Response(formatted_stats)

    @action(detail=False, methods=['get'], permission_classes=[IsStaffOrAdmin])
    def export_excel(self, request):
        """
        Exports all requests to an Excel spreadsheet.
        """
        queryset = self.get_queryset()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PIO Requests"
        
        # Define headers
        headers = [
            'Request ID', 'Agency/Requester', 'Contact Person', 
            'Barangay', 'Request Type', 'Status', 'Submitted At'
        ]
        
        # Style headers
        for col_num, column_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
        # Add data
        for row_num, obj in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1).value = obj.request_id
            sheet.cell(row=row_num, column=2).value = obj.requester.agency_name
            sheet.cell(row=row_num, column=3).value = obj.requester.contact_person
            sheet.cell(row=row_num, column=4).value = obj.requester.get_barangay_display()
            sheet.cell(row=row_num, column=5).value = obj.get_request_type_display()
            sheet.cell(row=row_num, column=6).value = obj.status.status_name if obj.status else 'Pending'
            sheet.cell(row=row_num, column=7).value = obj.submitted_at.strftime('%Y-%m-%d %H:%M')
            
        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=PIO_Requests_{timezone.now().strftime("%Y%m%d")}.xlsx'
        workbook.save(response)
        
        log_action(request.user, 'EXPORT EXCEL', 'Exported requests to Excel')
        return response

    @action(detail=True, methods=['get'], permission_classes=[IsOwnerOrStaff])
    def export_pdf(self, request, pk=None):
        """
        Generates a PDF report for a specific request.
        """
        obj = self.get_object()
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Header
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width/2, height - 50, "PUBLIC INFORMATION OFFICE (PIO)")
        p.setFont("Helvetica", 12)
        p.drawCentredString(width/2, height - 70, "Gumaca, Quezon, Philippines")
        p.line(50, height - 80, width - 50, height - 80)
        
        # Request Info
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, height - 110, f"OFFICIAL REQUEST REPORT #{obj.request_id}")
        
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, height - 140, "REQUESTER DETAILS")
        p.setFont("Helvetica", 11)
        p.drawString(60, height - 160, f"Agency/Name: {obj.requester.agency_name}")
        p.drawString(60, height - 175, f"Contact Person: {obj.requester.contact_person}")
        p.drawString(60, height - 190, f"Contact Number: {obj.requester.contact_number}")
        p.drawString(60, height - 205, f"Address: {obj.requester.address}, {obj.requester.get_barangay_display()}")
        
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, height - 235, "REQUEST INFORMATION")
        p.setFont("Helvetica", 11)
        p.drawString(60, height - 255, f"Type: {obj.get_request_type_display()}")
        p.drawString(60, height - 270, f"Status: {obj.status.status_name if obj.status else 'Pending'}")
        p.drawString(60, height - 285, f"Submitted At: {obj.submitted_at.strftime('%B %d, %Y %I:%M %p')}")
        
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, height - 315, "DETAILS / DESCRIPTION")
        p.setFont("Helvetica", 11)
        text_object = p.beginText(60, height - 335)
        text_object.setFont("Helvetica", 11)
        # Wrap text manually or use platypus for better wrapping
        details = obj.details or "No details provided."
        lines = [details[i:i+85] for i in range(0, len(details), 85)]
        for line in lines:
            text_object.textLine(line)
        p.drawText(text_object)
        
        # Approvals
        p.setFont("Helvetica-Bold", 11)
        curr_y = height - 450
        p.drawString(50, curr_y, "APPROVAL STATUS")
        curr_y -= 20
        approvals = obj.approvals.all()
        if approvals:
            for app in approvals:
                p.setFont("Helvetica", 11)
                p.drawString(60, curr_y, f"Status: {app.approval_status.upper()}")
                p.drawString(60, curr_y - 15, f"Remarks: {app.remarks or 'No remarks'}")
                p.drawString(60, curr_y - 30, f"Processed By: {app.approved_by.username if app.approved_by else 'N/A'}")
                curr_y -= 50
        else:
            p.drawString(60, curr_y, "Status: PENDING REVIEW")
            
        # Footer
        p.setFont("Helvetica-Oblique", 9)
        p.drawCentredString(width/2, 50, f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} - Archive-PIO System")
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        log_action(request.user, 'EXPORT PDF', f'Generated PDF for Request #{obj.request_id}')
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Request_{obj.request_id}_Report.pdf'
        return response


class ApprovalViewSet(viewsets.ModelViewSet):
    queryset = Approval.objects.all()
    serializer_class = ApprovalSerializer
    permission_classes = [IsStaffOrAdmin]

    def perform_create(self, serializer):
        instance = serializer.save(approved_by=self.request.user)
        log_action(self.request.user, 'CREATE APPROVAL',
                   f'Approval #{instance.approval_id}: {instance.approval_status}')

        # Find the citizen user who submitted the request (match by requester email)
        requester_email = instance.request.requester.email
        citizen = User.objects.filter(email=requester_email).first()

        if instance.approval_status == 'approved':
            message = (
                f'Good news! Your request #{instance.request.request_id} '
                f'({instance.request.get_request_type_display()}) '
                f'has been APPROVED.\n\nRemarks: {instance.remarks or "None"}'
            )
            notification_type = 'request_approved'
        elif instance.approval_status == 'rejected':
            message = (
                f'Your request #{instance.request.request_id} '
                f'({instance.request.get_request_type_display()}) '
                f'has been REJECTED.\n\nRemarks: {instance.remarks or "None"}'
            )
            notification_type = 'request_rejected'
        else:
            message = (
                f'Your request #{instance.request.request_id} '
                f'is now PENDING approval.'
            )
            notification_type = 'general'

        # Notify the citizen if found, otherwise notify the approver as fallback
        notify_user = citizen if citizen else self.request.user
        send_notification(
            user=notify_user,
            message=message,
            notification_type=notification_type
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(self.request.user, 'UPDATE APPROVAL',
                   f'Updated approval #{instance.approval_id}: {instance.approval_status}')


class AttachmentViewSet(viewsets.ModelViewSet):
    queryset = Attachment.objects.all()
    serializer_class = AttachmentSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def perform_create(self, serializer):
        uploader = self.request.user if self.request.user.is_authenticated else None
        instance = serializer.save(uploaded_by=uploader)
        if self.request.user.is_authenticated:
            log_action(self.request.user, 'UPLOAD ATTACHMENT',
                       f'Uploaded: {instance.file_name}')


class PublicationViewSet(viewsets.ModelViewSet):
    queryset = Publication.objects.all()
    serializer_class = PublicationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        instance = serializer.save(published_by=self.request.user)
        log_action(self.request.user, 'CREATE PUBLICATION',
                   f'Published to {instance.platform}')


class ProcessingLogViewSet(viewsets.ModelViewSet):
    queryset = ProcessingLog.objects.all()
    serializer_class = ProcessingLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        instance = serializer.save(processed_by=self.request.user)
        log_action(self.request.user, 'PROCESSING LOG',
                   f'Step: {instance.process_step}')