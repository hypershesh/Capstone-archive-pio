from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponse
from django.utils import timezone
from .models import (Requester, Status, Platform, Request,
                     Approval, Attachment, Publication, ProcessingLog)
import openpyxl
from openpyxl.styles import Font, Alignment

@admin.register(Request)
class RequestAdmin(admin.ModelAdmin):
    list_display = ('request_id', 'requester', 'request_type', 'status', 'submitted_at', 'download_pdf_link')
    list_filter = ('status', 'request_type', 'submitted_at')
    search_fields = ('requester__agency_name', 'details')
    actions = ['export_as_excel']

    def download_pdf_link(self, obj):
        return format_html(
            '<a class="button" style="background-color: #264b5d; color: white; padding: 2px 8px; border-radius: 4px; text-decoration: none;" href="{}" target="_blank">📄 PDF</a>',
            f'/api/requests/{obj.request_id}/export_pdf/'
        )
    download_pdf_link.short_description = 'Download PDF'

    @admin.action(description='Export Selected to Excel')
    def export_as_excel(self, request, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PIO Requests"
        
        headers = ['ID', 'Agency', 'Contact', 'Barangay', 'Type', 'Status', 'Date']
        for col_num, column_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            
        for row_num, obj in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1).value = obj.request_id
            sheet.cell(row=row_num, column=2).value = obj.requester.agency_name
            sheet.cell(row=row_num, column=3).value = obj.requester.contact_person
            sheet.cell(row=row_num, column=4).value = obj.requester.get_barangay_display()
            sheet.cell(row=row_num, column=5).value = obj.get_request_type_display()
            sheet.cell(row=row_num, column=6).value = obj.status.status_name if obj.status else 'Pending'
            sheet.cell(row=row_num, column=7).value = obj.submitted_at.strftime('%Y-%m-%d')
            
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=PIO_Requests_Export_{timezone.now().strftime("%Y%m%d")}.xlsx'
        workbook.save(response)
        return response

admin.site.register(Requester)
admin.site.register(Status)
admin.site.register(Platform)
admin.site.register(Approval)
admin.site.register(Attachment)
admin.site.register(Publication)
admin.site.register(ProcessingLog)
